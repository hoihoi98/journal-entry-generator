import io
import base64
from zipfile import ZipFile
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _detect_mime(data: bytes) -> str:
    if data[:4] == b'\x89PNG':
        return "image/png"
    if data[:3] == b'\xff\xd8\xff':
        return "image/jpeg"
    if data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        return "image/webp"
    return "image/png"


def _resolve_path(from_path: str, target: str) -> str:
    if target.startswith('/'):
        return target.lstrip('/')
    base_dir = from_path.rsplit('/', 1)[0] if '/' in from_path else ''
    raw = f"{base_dir}/{target}" if base_dir else target
    parts = raw.split('/')
    stack = []
    for p in parts:
        if p == '..':
            if stack:
                stack.pop()
        elif p and p != '.':
            stack.append(p)
    return '/'.join(stack)


def _rels_path(file_path: str) -> str:
    dir_part, name_part = file_path.rsplit('/', 1) if '/' in file_path else ('', file_path)
    return f"{dir_part}/_rels/{name_part}.rels" if dir_part else f"_rels/{name_part}.rels"


def extract_images_from_xlsx(xlsx_bytes: bytes) -> dict[str, list[tuple[bytes, str]]]:
    """Returns {sheet_name: [(image_bytes, mime_type), ...]} in workbook sheet order."""
    wb_ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    r_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    result: dict[str, list[tuple[bytes, str]]] = {}

    with ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        files = set(zf.namelist())

        wb_tree = ET.fromstring(zf.read('xl/workbook.xml'))
        # Preserve sheet order from the XML
        ordered_sheets = [
            (s.get(f'{{{r_ns}}}id'), s.get('name'))
            for s in wb_tree.iter(f'{{{wb_ns}}}sheet')
        ]

        rels_tree = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        rid_to_path = {
            rel.get('Id'): _resolve_path('xl/workbook.xml', rel.get('Target', ''))
            for rel in rels_tree
        }

        for rid, sheet_name in ordered_sheets:
            result[sheet_name] = []
            sheet_path = rid_to_path.get(rid)
            if not sheet_path:
                continue

            sheet_rels_path = _rels_path(sheet_path)
            if sheet_rels_path not in files:
                continue

            sheet_rels = ET.fromstring(zf.read(sheet_rels_path))
            for rel in sheet_rels:
                if 'drawing' not in rel.get('Type', '').lower():
                    continue
                drawing_path = _resolve_path(sheet_path, rel.get('Target', ''))
                if drawing_path not in files:
                    continue

                drawing_rels_path = _rels_path(drawing_path)
                if drawing_rels_path not in files:
                    continue

                drawing_rels = ET.fromstring(zf.read(drawing_rels_path))
                for drel in drawing_rels:
                    if 'image' not in drel.get('Type', '').lower():
                        continue
                    img_path = _resolve_path(drawing_path, drel.get('Target', ''))
                    if img_path in files:
                        data = zf.read(img_path)
                        result[sheet_name].append((data, _detect_mime(data)))

    return result


BATCH_PROMPT_ADDENDUM = """
BATCH PROCESSING INSTRUCTIONS
==============================
Your response MUST begin with these three lines, before the journal entry title:

DOCUMENT_NUMBERS: [all invoice/document reference numbers visible in the image(s), comma-separated — or "N/A" if none found]
TRANSACTION_CURRENCY: [ISO code of the primary currency of the transaction, e.g. EUR — or "N/A" if not identifiable]
TRANSACTION_AMOUNT: [total net amount of the transaction as a plain number, e.g. 1250.00 — or "N/A" if not identifiable]

Then output the journal entry in the standard format.
"""


def build_batch_user_message(
    images: list[tuple[bytes, str]],
    sheet_name: str,
    fx_line: str | None = None,
) -> list:
    content = [
        {
            "type": "image_url",
            "image_url": {"url": f"data:{mime};base64,{base64.b64encode(data).decode()}"},
        }
        for data, mime in images
    ]
    n = len(images)
    text = (
        f"This document belongs to transaction '{sheet_name}'. Produce the journal entry for it."
        if n == 1
        else f"These {n} documents all belong to transaction '{sheet_name}'. Produce ONE consolidated journal entry covering all of them."
    )
    if fx_line:
        text = f"{fx_line}\n{text}"
    content.append({"type": "text", "text": text})
    return content


def parse_batch_response(text: str) -> tuple[str, str, str, str]:
    """Returns (doc_numbers, txn_currency, txn_amount, journal_entry)."""
    doc_numbers = "N/A"
    txn_currency = ""
    txn_amount = ""
    journal_lines = []
    in_journal = False
    for line in text.strip().splitlines():
        if line.startswith("DOCUMENT_NUMBERS:"):
            doc_numbers = line[len("DOCUMENT_NUMBERS:"):].strip() or "N/A"
        elif line.startswith("TRANSACTION_CURRENCY:"):
            txn_currency = line[len("TRANSACTION_CURRENCY:"):].strip()
        elif line.startswith("TRANSACTION_AMOUNT:"):
            txn_amount = line[len("TRANSACTION_AMOUNT:"):].strip()
        else:
            if line.strip():
                in_journal = True
            if in_journal:
                journal_lines.append(line)
    return doc_numbers, txn_currency, txn_amount, "\n".join(journal_lines).strip()


def create_output_excel(results: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Entries"

    blue = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    headers = [
        "Transaction Number",
        "Invoice/Document Number",
        "Transaction Amount (Native)",
        "Transaction Amount (Functional)",
        "Journal Posting",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = blue
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=r["transaction_number"]).alignment = Alignment(vertical="top")
        ws.cell(row=i, column=2, value=r["doc_numbers"]).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=i, column=3, value=r.get("native_amount", "")).alignment = Alignment(vertical="top")
        ws.cell(row=i, column=4, value=r.get("functional_amount", "")).alignment = Alignment(vertical="top")
        entry = r.get("journal_entry") or r.get("error", "")
        cell = ws.cell(row=i, column=5, value=entry)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = max(40, (entry.count('\n') + 1) * 15)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
